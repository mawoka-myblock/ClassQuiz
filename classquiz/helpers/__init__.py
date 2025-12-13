# SPDX-FileCopyrightText: 2023 Marlon W (Mawoka)
#
# SPDX-License-Identifier: MPL-2.0


import asyncio
import uuid
from datetime import datetime
from typing import BinaryIO, Any

from fastapi import HTTPException
from openpyxl import load_workbook


from classquiz.db.models import Quiz, User, QuizQuestion, ABCDQuizAnswer
import xlsxwriter
from aiohttp import ClientSession
from io import BytesIO
from PIL import Image
from classquiz.config import meilisearch, settings, LOGGER
from classquiz.helpers.hashcash import check as hc_check

settings = settings()


async def get_meili_data(quiz: Quiz) -> dict[str, Any]:
    return {
        "id": str(quiz.id),
        "title": quiz.title,
        "description": quiz.description,
        "user": (await User.objects.filter(id=quiz.user_id).first()).username,
        "created_at": int(quiz.created_at.timestamp()),
        "imported_from_kahoot": quiz.imported_from_kahoot,
    }


async def generate_spreadsheet(
    quiz_results: dict[str, Any],
    quiz: Quiz,
    player_fields: dict[str, Any],
    player_scores: dict[str, Any],
) -> BytesIO:
    storage = BytesIO()
    workbook = xlsxwriter.Workbook(storage, {"in_memory": True})
    player_worksheet = workbook.add_worksheet()
    player_worksheet.name = "Players"
    _ = player_worksheet.write(0, 0, "Username")
    _ = player_worksheet.write(0, 1, "Score")
    _ = player_worksheet.write(0, 2, "Custom-Field")
    for i, player in enumerate(player_scores.keys()):
        player_worksheet.write(i + 1, 0, player)
        player_worksheet.write(i + 1, 1, player_scores[player])
        try:
            player_worksheet.write(i + 1, 2, player_fields[player])
        except KeyError:
            continue

    worksheet = workbook.add_worksheet()
    worksheet.name = "Questions"
    _ = worksheet.write(0, 0, "Question")
    _ = worksheet.write(0, 1, "Time")
    _ = worksheet.write(0, 2, "Image")
    _ = worksheet.write(0, 3, "Correct answers")
    _ = worksheet.write(0, 4, "Correct answers")
    _ = worksheet.write(0, 5, "Wrong answers")
    for i, _ in enumerate(quiz_results):
        question = quiz.questions[i]
        try:
            answer_data = quiz_results[str(i)]
        except KeyError:
            continue
        _ = worksheet.write(i + 1, 0, question["question"])
        _ = worksheet.write(i + 1, 1, question["time"])

        try:
            async with (
                ClientSession() as session,
                session.get(f"{settings.root_address}/api/v1/storage/download/{question['image']}") as response,
            ):
                content_type = response.headers.get("Content-Type")
                if content_type is not None and "image" in content_type:
                    img_data = BytesIO(await response.read())
                    _ = worksheet.insert_image(i + 1, 2, question["image"], {"image_data": img_data})
                    image = Image.open(img_data)
                    _ = worksheet.set_row(i + 1, image.height)
                    _ = worksheet.set_column(2, 2, image.width)
        except TypeError:
            pass
        answer_amount = len(answer_data)
        correct_answers = 0
        wrong_answers = 0
        for j in answer_data:
            j_correct = j["right"]
            if j_correct:
                correct_answers += 1
            else:
                wrong_answers += 1
        _ = worksheet.write(i + 1, 3, f"{round(correct_answers / answer_amount * 100)}%")
        _ = worksheet.write(i + 1, 4, correct_answers)
        _ = worksheet.write(i + 1, 5, wrong_answers)

        ws = workbook.add_worksheet(f"{i + 1}. Question")
        _ = ws.write(0, 0, "Answer")
        _ = ws.write(0, 1, "Correct")
        _ = ws.write(0, 2, "Username")
        for j, _ in enumerate(answer_data):
            _ = ws.write(j + 1, 0, answer_data[j]["answer"])
            if answer_data[j]["right"]:
                _ = ws.write(j + 1, 1, "True")
            else:
                _ = ws.write(j + 1, 1, "False")
            _ = ws.write(j + 1, 2, answer_data[j]["username"])

    workbook.close()
    _ = storage.seek(0)
    return storage


async def handle_import_from_excel(data: BinaryIO, user: User) -> Quiz:
    try:
        wb = load_workbook(filename=data)
    except KeyError:
        raise HTTPException(status_code=400, detail="File not in excel format")
    ws = wb.active
    if ws is None:
        raise Exception("Workbook is None")
    questions: list[dict] = []
    title: str | None = ws["C5"].value
    description: str | None = ws["C6"].value
    if title is None:
        raise HTTPException(status_code=400, detail="Title missing")
    if description is None:
        raise HTTPException(status_code=400, detail="Description missing")
    if not 3 <= len(description) <= 500:
        raise HTTPException(status_code=400, detail="Description doesn't have the correct length")
    if not 3 <= len(title) <= 500:
        raise HTTPException(status_code=400, detail="Title doesn't have the correct length")
    existing_quiz: Quiz | None = await Quiz.objects.get_or_none(user_id=user, title=title)
    for i, row in enumerate(ws.iter_rows(min_row=14, min_col=2, max_row=100, max_col=8, values_only=True)):
        if row[0] is None:
            continue
        question = str(row[0])
        if len(question) > 300:
            raise HTTPException(status_code=400, detail=f"Question {i + 1} is longer than 300")
        try:
            time = int(str(row[5]).partition(".")[0])
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail=f"Time in q {i + 1} isn't a valid int")
        if time > 120:
            raise HTTPException(status_code=400, detail=f"Time in q {i + 1} is greater than 120")
        answers = [row[1], row[2], row[3], row[4]]
        for a, answer in enumerate(answers):
            if answer is None:
                continue
            if len(str(answer)) > 150:
                raise HTTPException(
                    status_code=400,
                    detail=f"Answer {a + 1} in Question {i + 1} is longer than 150",
                )
        if len(answers) < 2:
            raise HTTPException(status_code=400, detail=f"Less than 2 answers in Question {i + 1}")
        correct_answers: str = str(row[6])
        answers_list: list[ABCDQuizAnswer] = []
        for a, answer in enumerate(answers):
            if answer is None:
                continue
            answers_list.append(ABCDQuizAnswer(answer=str(answer), right=str(a + 1) in correct_answers))

        existing_question: dict | None = None
        if existing_quiz is not None:
            existing_question: dict = existing_quiz.questions[i]
        questions.append(
            QuizQuestion(
                question=question,
                answers=answers_list,
                time=str(time),
                image=existing_question["image"] if existing_question is not None else None,
            ).model_dump()
        )
    if existing_quiz is None:
        quiz = Quiz(
            title=title,
            description=description,
            id=uuid.uuid4(),
            created_at=datetime.now(),
            updated_at=datetime.now(),
            user_id=user,
            questions=questions,
            imported_from_kahoot=False,
        )
        _ = await quiz.save()
    else:
        existing_quiz.questions = [*existing_quiz.questions, *questions]
        existing_quiz.updated_at = datetime.now()
        existing_quiz.mod_rating = None
        _ = await existing_quiz.update()
        quiz = existing_quiz
    return quiz


async def meilisearch_init():
    classquiz_index_found = False
    try:
        indexes = meilisearch.get_indexes()
        for index in indexes["results"]:
            if index.uid == settings.meilisearch_index:
                classquiz_index_found = True
    except TypeError:
        classquiz_index_found = False
    # +++ Check if the index does not exist and creates it
    if not classquiz_index_found:
        LOGGER.debug("Creating MeiliSearch Index")
        meilisearch.create_index(settings.meilisearch_index)
        await asyncio.sleep(1)
    # --- END
    # +++ Check if count of docs in meilisearch is equal the count in the database +++
    quiz_count = await Quiz.objects.filter(public=True).count()
    number_of_docs = 0
    for i in meilisearch.index(settings.meilisearch_index).get_stats():
        number_of_docs = i[1]
        break
    if number_of_docs != quiz_count:
        LOGGER.warn("MeiliSearch and Database got out of sync, syncthing them")
        meilisearch.delete_index(settings.meilisearch_index)
        meilisearch.create_index(settings.meilisearch_index)
        quizzes = await Quiz.objects.filter(public=True).all()
        meili_data = []
        for quiz in quizzes:
            meili_data.append(await get_meili_data(quiz))
        meilisearch.index(settings.meilisearch_index).add_documents(meili_data)
    # --- END ---
    meilisearch.index(settings.meilisearch_index).update_settings({"sortableAttributes": ["created_at"]})
    LOGGER.info("Finished MeiliSearch synchronisation")


def check_hashcash(data: str, input_data: str, claim_in: str | None = "19") -> bool:
    """
    It checks that the hashcash is valid, and if it is, it returns True

    :param claim_in: The claim the data should have (bytes)
    :param data: The hashcash string
    :type data: str
    :param input_data: The claim, the data the server provided
    :type input_data: str
    :return: A boolean value.
    """

    if not hc_check(data):
        return False
    try:
        version, claim, _date, res, ext, _rand, _counter = data.split(":")
    except ValueError:
        return False
    some_error = [version == "1", claim == claim_in, res == input_data, ext == ""]
    return all(el is True for el in some_error)


def check_image_string(image: str) -> (bool, uuid.UUID | None):
    # Valid formats: {uuid} and {uuid}--{uuid}
    try:
        parsed_uuid = uuid.UUID(image)
        return True, parsed_uuid
    except ValueError:
        pass

    split_image = image.split("--")
    if len(split_image) != 2:
        return False, None

    try:
        uuid.UUID(split_image[0])
        uuid.UUID(split_image[1])
        return True, None
    except ValueError:
        return False, None


def extract_image_ids_from_quiz(quiz: Quiz) -> list[str | uuid.UUID]:
    quiz_images = []
    if quiz.background_image is not None:
        quiz_images.append(quiz.background_image)
    if quiz.cover_image is not None:
        quiz_images.append(quiz.cover_image)
    for question in quiz.questions:
        if question["image"] is None:
            continue
        quiz_images.append(question["image"])
    return quiz_images
